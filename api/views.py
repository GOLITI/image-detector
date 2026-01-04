"""
ViewSets pour l'API REST
"""
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter
from datetime import timedelta
import csv
import json
from io import BytesIO

from .models import User, Category, Tag, BatchAnalysis
from .serializers import (
    UserRegistrationSerializer, UserLoginSerializer, UserSerializer,
    UserProfileUpdateSerializer, ChangePasswordSerializer, ApiKeySerializer,
    CategorySerializer, TagSerializer,
    ImageAnalysisListSerializer, ImageAnalysisDetailSerializer,
    ImageAnalysisCreateSerializer, ImageAnalysisUpdateSerializer,
    BatchAnalysisListSerializer, BatchAnalysisDetailSerializer,
    BatchAnalysisCreateSerializer, ExportRequestSerializer,
    UserStatisticsSerializer
)
from detector.models import ImageAnalysis
from services.forgery_detector import ForgeryDetector
from services.report_generator import ReportGenerator


# ============================================
# AUTHENTICATION VIEWS
# ============================================

@extend_schema_view(
    create=extend_schema(description="Inscription d'un nouvel utilisateur", tags=['auth']),
)
class AuthViewSet(viewsets.ViewSet):
    """ViewSet pour l'authentification"""
    permission_classes = [permissions.AllowAny]
    
    @extend_schema(
        request=UserRegistrationSerializer,
        responses={201: UserSerializer},
        tags=['auth']
    )
    @action(detail=False, methods=['post'])
    def register(self, request):
        """Inscription d'un nouvel utilisateur"""
        serializer = UserRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            token, _ = Token.objects.get_or_create(user=user)
            return Response({
                'user': UserSerializer(user).data,
                'token': token.key,
                'message': 'Inscription réussie!'
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @extend_schema(
        request=UserLoginSerializer,
        responses={200: UserSerializer},
        tags=['auth']
    )
    @action(detail=False, methods=['post'])
    def login(self, request):
        """Connexion utilisateur"""
        serializer = UserLoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']
            token, _ = Token.objects.get_or_create(user=user)
            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])
            return Response({
                'user': UserSerializer(user).data,
                'token': token.key,
                'message': 'Connexion réussie!'
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @extend_schema(tags=['auth'])
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def logout(self, request):
        """Déconnexion - supprime le token"""
        try:
            request.user.auth_token.delete()
        except:
            pass
        return Response({'message': 'Déconnexion réussie!'})


# ============================================
# USER VIEWSET
# ============================================

@extend_schema_view(
    retrieve=extend_schema(description="Récupérer les informations d'un utilisateur", tags=['auth']),
    update=extend_schema(description="Mettre à jour le profil", tags=['auth']),
    partial_update=extend_schema(description="Mise à jour partielle du profil", tags=['auth']),
)
class UserViewSet(viewsets.ModelViewSet):
    """ViewSet pour la gestion des utilisateurs"""
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return User.objects.filter(id=self.request.user.id)
    
    def get_object(self):
        return self.request.user
    
    @extend_schema(tags=['auth'])
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Récupérer le profil de l'utilisateur connecté"""
        serializer = UserSerializer(request.user)
        return Response(serializer.data)
    
    @extend_schema(request=UserProfileUpdateSerializer, tags=['auth'])
    @action(detail=False, methods=['patch'])
    def update_profile(self, request):
        """Mettre à jour le profil"""
        serializer = UserProfileUpdateSerializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(UserSerializer(request.user).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @extend_schema(request=ChangePasswordSerializer, tags=['auth'])
    @action(detail=False, methods=['post'])
    def change_password(self, request):
        """Changer le mot de passe"""
        serializer = ChangePasswordSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            request.user.set_password(serializer.validated_data['new_password'])
            request.user.save()
            return Response({'message': 'Mot de passe modifié avec succès!'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @extend_schema(responses={200: ApiKeySerializer}, tags=['auth'])
    @action(detail=False, methods=['post'])
    def generate_api_key(self, request):
        """Générer une nouvelle clé API"""
        api_key = request.user.generate_api_key()
        return Response({
            'api_key': api_key,
            'message': 'Nouvelle clé API générée. Conservez-la précieusement!'
        })
    
    @extend_schema(responses={200: UserStatisticsSerializer}, tags=['auth'])
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Récupérer les statistiques de l'utilisateur"""
        user = request.user
        now = timezone.now()
        
        analyses = ImageAnalysis.objects.filter(user=user)
        
        # Distribution des verdicts
        verdict_dist = dict(analyses.values('verdict').annotate(count=Count('id')).values_list('verdict', 'count'))
        
        # Catégorie la plus utilisée
        most_used_cat = analyses.exclude(category__isnull=True).values('category__name').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        # Tags les plus utilisés
        most_used_tags = list(
            Tag.objects.filter(analyses__user=user).annotate(
                count=Count('analyses')
            ).order_by('-count').values_list('name', flat=True)[:5]
        )
        
        stats = {
            'total_analyses': analyses.count(),
            'analyses_this_month': analyses.filter(created_at__gte=now - timedelta(days=30)).count(),
            'analyses_this_week': analyses.filter(created_at__gte=now - timedelta(days=7)).count(),
            'verdict_distribution': verdict_dist,
            'average_similarity': analyses.aggregate(avg=Avg('similarity_percentage'))['avg'] or 0,
            'average_analysis_duration': analyses.aggregate(avg=Avg('analysis_duration'))['avg'] or 0,
            'total_categories': user.categories.count(),
            'total_tags': user.tags.count(),
            'total_batches': user.batch_analyses.count(),
            'most_used_category': most_used_cat['category__name'] if most_used_cat else None,
            'most_used_tags': most_used_tags,
        }
        
        return Response(UserStatisticsSerializer(stats).data)


# ============================================
# CATEGORY & TAG VIEWSETS
# ============================================

@extend_schema_view(
    list=extend_schema(description="Liste des catégories", tags=['categories']),
    create=extend_schema(description="Créer une catégorie", tags=['categories']),
    retrieve=extend_schema(description="Détail d'une catégorie", tags=['categories']),
    update=extend_schema(description="Modifier une catégorie", tags=['categories']),
    destroy=extend_schema(description="Supprimer une catégorie", tags=['categories']),
)
class CategoryViewSet(viewsets.ModelViewSet):
    """ViewSet pour les catégories"""
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    
    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)


@extend_schema_view(
    list=extend_schema(description="Liste des tags", tags=['tags']),
    create=extend_schema(description="Créer un tag", tags=['tags']),
    retrieve=extend_schema(description="Détail d'un tag", tags=['tags']),
    update=extend_schema(description="Modifier un tag", tags=['tags']),
    destroy=extend_schema(description="Supprimer un tag", tags=['tags']),
)
class TagViewSet(viewsets.ModelViewSet):
    """ViewSet pour les tags"""
    serializer_class = TagSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [SearchFilter]
    search_fields = ['name']
    
    def get_queryset(self):
        return Tag.objects.filter(user=self.request.user)


# ============================================
# IMAGE ANALYSIS VIEWSET
# ============================================

@extend_schema_view(
    list=extend_schema(description="Liste des analyses de l'utilisateur", tags=['analyses']),
    create=extend_schema(description="Créer une nouvelle analyse", tags=['analyses']),
    retrieve=extend_schema(description="Détail d'une analyse", tags=['analyses']),
    update=extend_schema(description="Modifier une analyse", tags=['analyses']),
    destroy=extend_schema(description="Supprimer une analyse", tags=['analyses']),
)
class ImageAnalysisViewSet(viewsets.ModelViewSet):
    """ViewSet pour les analyses d'images"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['verdict', 'category', 'is_favorite']
    search_fields = ['title', 'notes']
    ordering_fields = ['created_at', 'similarity_percentage', 'ssim_score']
    ordering = ['-created_at']
    
    def get_queryset(self):
        return ImageAnalysis.objects.filter(user=self.request.user).select_related('category').prefetch_related('tags')
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ImageAnalysisListSerializer
        elif self.action == 'create':
            return ImageAnalysisCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return ImageAnalysisUpdateSerializer
        return ImageAnalysisDetailSerializer
    
    def create(self, request, *args, **kwargs):
        """Créer et exécuter une analyse"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Créer l'instance
        analysis = ImageAnalysis(
            image1=serializer.validated_data['image1'],
            image2=serializer.validated_data['image2'],
            title=serializer.validated_data.get('title', ''),
            notes=serializer.validated_data.get('notes', ''),
            user=request.user
        )
        
        # Catégorie
        category_id = serializer.validated_data.get('category_id')
        if category_id:
            analysis.category_id = category_id
        
        analysis.save()
        
        # Tags
        tag_ids = serializer.validated_data.get('tag_ids', [])
        if tag_ids:
            analysis.tags.set(tag_ids)
        
        # Exécuter l'analyse
        try:
            analysis = ForgeryDetector.analyze_images(analysis)
            analysis.save()
            
            # Mettre à jour les statistiques
            request.user.increment_analyses()
            
            return Response(
                ImageAnalysisDetailSerializer(analysis, context={'request': request}).data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            analysis.delete()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @extend_schema(tags=['analyses'])
    @action(detail=True, methods=['get'])
    def report(self, request, pk=None):
        """Télécharger le rapport PDF"""
        analysis = self.get_object()
        
        try:
            pdf_buffer = ReportGenerator.generate_report(analysis)
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="rapport_analyse_{analysis.id}.pdf"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @extend_schema(tags=['analyses'])
    @action(detail=True, methods=['post'])
    def toggle_favorite(self, request, pk=None):
        """Ajouter/retirer des favoris"""
        analysis = self.get_object()
        analysis.is_favorite = not analysis.is_favorite
        analysis.save(update_fields=['is_favorite'])
        return Response({'is_favorite': analysis.is_favorite})
    
    @extend_schema(tags=['analyses'])
    @action(detail=False, methods=['get'])
    def favorites(self, request):
        """Liste des analyses favorites"""
        queryset = self.get_queryset().filter(is_favorite=True)
        serializer = ImageAnalysisListSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(tags=['analyses'])
    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Analyses récentes (7 derniers jours)"""
        week_ago = timezone.now() - timedelta(days=7)
        queryset = self.get_queryset().filter(created_at__gte=week_ago)
        serializer = ImageAnalysisListSerializer(queryset, many=True)
        return Response(serializer.data)


# ============================================
# BATCH ANALYSIS VIEWSET
# ============================================

@extend_schema_view(
    list=extend_schema(description="Liste des analyses par lot", tags=['batch']),
    create=extend_schema(description="Créer une analyse par lot", tags=['batch']),
    retrieve=extend_schema(description="Détail d'un lot", tags=['batch']),
    destroy=extend_schema(description="Supprimer un lot", tags=['batch']),
)
class BatchAnalysisViewSet(viewsets.ModelViewSet):
    """ViewSet pour les analyses par lot"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status']
    ordering = ['-created_at']
    
    def get_queryset(self):
        return BatchAnalysis.objects.filter(user=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'list':
            return BatchAnalysisListSerializer
        elif self.action == 'create':
            return BatchAnalysisCreateSerializer
        return BatchAnalysisDetailSerializer
    
    @extend_schema(request=BatchAnalysisCreateSerializer, tags=['batch'])
    def create(self, request, *args, **kwargs):
        """Créer et exécuter une analyse par lot"""
        # Récupérer les données
        name = request.data.get('name', f'Lot du {timezone.now().strftime("%d/%m/%Y %H:%M")}')
        description = request.data.get('description', '')
        
        # Récupérer les paires d'images
        pairs = []
        i = 0
        while f'pairs[{i}][image1]' in request.FILES or f'image1_{i}' in request.FILES:
            image1_key = f'pairs[{i}][image1]' if f'pairs[{i}][image1]' in request.FILES else f'image1_{i}'
            image2_key = f'pairs[{i}][image2]' if f'pairs[{i}][image2]' in request.FILES else f'image2_{i}'
            title_key = f'pairs[{i}][title]' if f'pairs[{i}][title]' in request.data else f'title_{i}'
            
            if image1_key in request.FILES and image2_key in request.FILES:
                pairs.append({
                    'image1': request.FILES[image1_key],
                    'image2': request.FILES[image2_key],
                    'title': request.data.get(title_key, f'Paire {i+1}')
                })
            i += 1
        
        if not pairs:
            return Response(
                {'error': 'Aucune paire d\'images fournie'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(pairs) > 50:
            return Response(
                {'error': 'Maximum 50 paires par lot'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Créer le lot
        batch = BatchAnalysis.objects.create(
            name=name,
            description=description,
            user=request.user,
            total_pairs=len(pairs),
            status='PROCESSING',
            started_at=timezone.now()
        )
        
        # Traiter chaque paire
        for pair in pairs:
            try:
                analysis = ImageAnalysis(
                    image1=pair['image1'],
                    image2=pair['image2'],
                    title=pair['title'],
                    user=request.user,
                    batch=batch
                )
                analysis.save()
                
                # Exécuter l'analyse
                analysis = ForgeryDetector.analyze_images(analysis)
                analysis.save()
                
                batch.successful_pairs += 1
            except Exception as e:
                batch.failed_pairs += 1
            
            batch.processed_pairs += 1
            batch.save()
        
        # Mettre à jour les statistiques finales
        batch.update_statistics()
        
        return Response(
            BatchAnalysisDetailSerializer(batch).data,
            status=status.HTTP_201_CREATED
        )


# ============================================
# EXPORT VIEWSET
# ============================================

@extend_schema_view(
    create=extend_schema(description="Exporter les données", tags=['export']),
)
class ExportViewSet(viewsets.ViewSet):
    """ViewSet pour l'export de données"""
    permission_classes = [permissions.IsAuthenticated]
    
    @extend_schema(
        request=ExportRequestSerializer,
        tags=['export'],
        parameters=[
            OpenApiParameter(name='format', description='Format d\'export (csv, excel, json)', required=False, type=str),
            OpenApiParameter(name='start_date', description='Date de début (YYYY-MM-DD)', required=False, type=str),
            OpenApiParameter(name='end_date', description='Date de fin (YYYY-MM-DD)', required=False, type=str),
        ]
    )
    @action(detail=False, methods=['get', 'post'])
    def analyses(self, request):
        """Exporter les analyses"""
        # Récupérer les paramètres
        export_format = request.data.get('format') or request.query_params.get('format', 'csv')
        start_date = request.data.get('start_date') or request.query_params.get('start_date')
        end_date = request.data.get('end_date') or request.query_params.get('end_date')
        verdict = request.data.get('verdict') or request.query_params.get('verdict')
        category_id = request.data.get('category_id') or request.query_params.get('category_id')
        
        # Filtrer les analyses
        queryset = ImageAnalysis.objects.filter(user=request.user)
        
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
        if verdict:
            queryset = queryset.filter(verdict=verdict)
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        analyses = queryset.order_by('-created_at')
        
        if export_format == 'csv':
            return self._export_csv(analyses)
        elif export_format == 'excel':
            return self._export_excel(analyses)
        elif export_format == 'json':
            return self._export_json(analyses)
        else:
            return Response({'error': 'Format non supporté'}, status=status.HTTP_400_BAD_REQUEST)
    
    def _export_csv(self, analyses):
        """Export en CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="analyses_{timezone.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')  # BOM pour Excel
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'ID', 'Titre', 'Date', 'Verdict', 'Similarité (%)', 'Score SSIM',
            'Hash MD5 Image 1', 'Hash MD5 Image 2', 'Durée (s)', 'Catégorie', 'Favori'
        ])
        
        for a in analyses:
            writer.writerow([
                a.id,
                a.title,
                a.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                a.get_verdict_display(),
                round(a.similarity_percentage, 2) if a.similarity_percentage else '',
                round(a.ssim_score, 4) if a.ssim_score else '',
                a.md5_hash1,
                a.md5_hash2,
                a.analysis_duration,
                a.category.name if a.category else '',
                'Oui' if a.is_favorite else 'Non'
            ])
        
        return response
    
    def _export_excel(self, analyses):
        """Export en Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl non installé'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analyses"
        
        # En-têtes
        headers = [
            'ID', 'Titre', 'Date', 'Verdict', 'Similarité (%)', 'Score SSIM',
            'Hash MD5 Image 1', 'Hash MD5 Image 2', 'Durée (s)', 'Catégorie', 'Favori'
        ]
        
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row, a in enumerate(analyses, 2):
            ws.cell(row=row, column=1, value=a.id)
            ws.cell(row=row, column=2, value=a.title)
            ws.cell(row=row, column=3, value=a.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=4, value=a.get_verdict_display())
            ws.cell(row=row, column=5, value=round(a.similarity_percentage, 2) if a.similarity_percentage else None)
            ws.cell(row=row, column=6, value=round(a.ssim_score, 4) if a.ssim_score else None)
            ws.cell(row=row, column=7, value=a.md5_hash1)
            ws.cell(row=row, column=8, value=a.md5_hash2)
            ws.cell(row=row, column=9, value=a.analysis_duration)
            ws.cell(row=row, column=10, value=a.category.name if a.category else '')
            ws.cell(row=row, column=11, value='Oui' if a.is_favorite else 'Non')
        
        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Sauvegarder
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="analyses_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    def _export_json(self, analyses):
        """Export en JSON"""
        data = []
        for a in analyses:
            data.append({
                'id': a.id,
                'title': a.title,
                'created_at': a.created_at.isoformat(),
                'verdict': a.verdict,
                'verdict_display': a.get_verdict_display(),
                'similarity_percentage': a.similarity_percentage,
                'ssim_score': a.ssim_score,
                'md5_hash1': a.md5_hash1,
                'md5_hash2': a.md5_hash2,
                'analysis_duration': a.analysis_duration,
                'category': a.category.name if a.category else None,
                'is_favorite': a.is_favorite,
                'ai_confidence_score': a.ai_confidence_score,
            })
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="analyses_{timezone.now().strftime("%Y%m%d")}.json"'
        return response
